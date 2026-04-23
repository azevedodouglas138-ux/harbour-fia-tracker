"""
research_excel_parser.py — Heurística Python para extrair campos de valuation
de modelos Excel (.xlsx / .xlsm).

Estratégia híbrida: este módulo é a primeira camada. Se os campos obrigatórios
não forem encontrados (REQUIRED_FIELDS), o frontend pode disparar uma extração
via Claude Haiku como fallback on-demand.

API pública:
  parse_excel(path) -> dict com {extracted, missing, scenarios, sensitivity, sheets}
  excel_to_markdown(path, max_cells=2000) -> str (input pro Claude)
  REQUIRED_FIELDS -> list[str]
"""

import logging
import re
from typing import Any, Optional

import openpyxl

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Label map — sinônimos bilíngues pt-BR / en
# ---------------------------------------------------------------------------

LABEL_MAP = {
    "target_price": [
        "preço alvo", "preco alvo", "preço-alvo", "preco-alvo",
        "target price", "fair value", "valor justo",
        "preço teto", "preco teto", "preço-teto",
        "pt", "target", "upside target",
    ],
    "upside": [
        "upside", "potencial", "retorno esperado",
        "upside (%)", "potencial de alta",
    ],
    "wacc": [
        "wacc", "taxa de desconto", "discount rate",
        "custo de capital", "custo medio ponderado", "custo médio ponderado",
        "ke", "cost of equity",
    ],
    "growth_rate": [
        "growth", "crescimento", "taxa de crescimento",
        "cagr", "revenue growth",
    ],
    "terminal_growth": [
        "perpetuidade", "terminal growth", "g perpetuidade",
        "taxa perpetua", "taxa perpétua", "crescimento perpetuo",
        "crescimento perpétuo", "g perp", "g terminal",
    ],
    "ebitda_margin": [
        "margem ebitda", "ebitda margin", "margem operacional",
        "ebitda/receita",
    ],
    "revenue_cagr": [
        "revenue cagr", "cagr receita", "cagr da receita",
        "crescimento receita", "receita cagr",
    ],
    "methodology": [
        "metodologia", "methodology", "metodo de valuation", "método de valuation",
    ],
}

REQUIRED_FIELDS = ["target_price", "wacc", "growth_rate"]

# Scenarios detection — labels bilíngues
_SCENARIO_LABELS = {
    "bear": ["bear", "pessimista", "conservador", "downside", "stress"],
    "base": ["base", "central", "provável", "provavel", "neutral"],
    "bull": ["bull", "otimista", "upside", "agressivo"],
}

_METHODOLOGY_KEYWORDS = {
    "DCF": ["dcf", "fluxo de caixa descontado", "discounted cash flow"],
    "EV/EBITDA": ["ev/ebitda", "ev ebitda", "múltiplo ebitda", "multiplo ebitda"],
    "P/L": ["p/l", "p/e", "preço/lucro", "preco/lucro"],
    "DDM": ["ddm", "dividend discount", "modelo de dividendos"],
    "SOMA_PARTES": ["soma das partes", "soma-das-partes", "sum of parts", "sotp"],
}


# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------

def _normalize(s: Any) -> str:
    """Lowercase, strip, collapse whitespace. Returns '' for non-strings."""
    if s is None:
        return ""
    s = str(s).lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _is_numeric(value: Any) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value.replace(",", ".").replace("%", "").strip())
            return True
        except (ValueError, TypeError):
            return False
    return False


def _to_number(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", ".").replace("%", "").strip())
        except (ValueError, TypeError):
            return None
    return None


def _is_percentage_cell(cell) -> bool:
    """Heuristic: True if cell is formatted as percentage OR value is in [0,1]."""
    fmt = (cell.number_format or "") if cell is not None else ""
    if "%" in fmt:
        return True
    val = _to_number(cell.value if cell is not None else None)
    if val is not None and 0 < val <= 1:
        return True
    return False


def _coerce_percent(cell) -> Optional[float]:
    """Extract numeric value from a cell, multiplying by 100 if fmt is percentage
    or raw value is fraction."""
    val = _to_number(cell.value)
    if val is None:
        return None
    # If format is % and value is 0.118, openpyxl returns 0.118 raw — multiply.
    fmt = (cell.number_format or "")
    if "%" in fmt:
        return round(val * 100, 4)
    # If no format but value is fraction-ish, assume it's already a decimal.
    if 0 < val <= 1:
        return round(val * 100, 4)
    return round(val, 4)


def _find_adjacent_numeric(sheet, row: int, col: int, max_steps: int = 3):
    """Look to the right first (up to max_steps cells), then down (up to max_steps).

    Returns (value, source_cell) or (None, None).
    """
    # Right
    for dc in range(1, max_steps + 1):
        c = col + dc
        if c > sheet.max_column:
            break
        cell = sheet.cell(row=row, column=c)
        if _is_numeric(cell.value):
            return cell.value, cell
    # Down
    for dr in range(1, max_steps + 1):
        r = row + dr
        if r > sheet.max_row:
            break
        cell = sheet.cell(row=r, column=col)
        if _is_numeric(cell.value):
            return cell.value, cell
    return None, None


# ---------------------------------------------------------------------------
# Scenarios detection
# ---------------------------------------------------------------------------

def _detect_scenarios(sheet) -> Optional[dict]:
    """Look for BEAR/BASE/BULL (or PT pessimista/base/otimista) headers in a
    row or column, then extract the 3 target prices adjacent to them.
    """
    max_r = min(sheet.max_row, 200)
    max_c = min(sheet.max_column, 50)

    # Scan cells
    for row in range(1, max_r + 1):
        for col in range(1, max_c + 1):
            cell = sheet.cell(row=row, column=col)
            norm = _normalize(cell.value)
            if not norm:
                continue
            # Check if cell label matches one of the scenario types
            for scen, synonyms in _SCENARIO_LABELS.items():
                if any(norm == s or f"cenario {s}" == norm or norm.startswith(s + " ") for s in synonyms):
                    # Found scenario label. Try to find the other two adjacent.
                    result = _collect_scenario_block(sheet, row, col)
                    if result and len(result) >= 2:  # at least 2 scenarios found
                        return result
    return None


def _collect_scenario_block(sheet, row: int, col: int) -> Optional[dict]:
    """Given a cell with a scenario label, try to find the other two along
    the same row or column and collect their numeric values."""
    out = {}

    # Try HORIZONTAL: scenarios in a row, values in next row
    labels_in_row = []
    for c in range(max(1, col - 3), min(sheet.max_column, col + 5) + 1):
        cell = sheet.cell(row=row, column=c)
        norm = _normalize(cell.value)
        for scen, synonyms in _SCENARIO_LABELS.items():
            if any(norm == s or norm.startswith(s) for s in synonyms):
                labels_in_row.append((c, scen))

    if len(labels_in_row) >= 2:
        # Look for values in next 1-3 rows
        for dr in range(1, 4):
            r = row + dr
            if r > sheet.max_row:
                break
            row_out = {}
            for c, scen in labels_in_row:
                val = _to_number(sheet.cell(row=r, column=c).value)
                if val is not None:
                    row_out[scen] = {"price": round(val, 4), "upside": None}
            if len(row_out) >= len(labels_in_row):
                return row_out

    # Try VERTICAL: scenarios in a column, values in adjacent column
    labels_in_col = []
    for r in range(max(1, row - 3), min(sheet.max_row, row + 5) + 1):
        cell = sheet.cell(row=r, column=col)
        norm = _normalize(cell.value)
        for scen, synonyms in _SCENARIO_LABELS.items():
            if any(norm == s or norm.startswith(s) for s in synonyms):
                labels_in_col.append((r, scen))

    if len(labels_in_col) >= 2:
        for dc in range(1, 4):
            c = col + dc
            if c > sheet.max_column:
                break
            col_out = {}
            for r, scen in labels_in_col:
                val = _to_number(sheet.cell(row=r, column=c).value)
                if val is not None:
                    col_out[scen] = {"price": round(val, 4), "upside": None}
            if len(col_out) >= len(labels_in_col):
                return col_out

    return None if not out else out


# ---------------------------------------------------------------------------
# Sensitivity matrix detection
# ---------------------------------------------------------------------------

def _detect_sensitivity(sheet) -> Optional[dict]:
    """Look for a 2D block where first column has WACC-like labels and first
    row has growth-like labels (or vice versa). Returns
    {rows: [...], cols: [...], matrix: [[...]], base: [r,c] or None}.
    """
    max_r = min(sheet.max_row, 300)
    max_c = min(sheet.max_column, 50)

    # Strategy: find a cell whose value matches "WACC" or "Taxa de Desconto"
    # and check if the row/col below/right has a matrix.
    for row in range(1, max_r + 1):
        for col in range(1, max_c + 1):
            cell = sheet.cell(row=row, column=col)
            norm = _normalize(cell.value)
            if not norm:
                continue
            if any(w in norm for w in ["wacc", "taxa de desconto"]):
                # Try block starting at (row, col)
                block = _try_extract_matrix(sheet, row, col)
                if block:
                    return block
    return None


def _try_extract_matrix(sheet, anchor_row: int, anchor_col: int):
    """Try to extract a sensitivity matrix anchored at `anchor_row,anchor_col`.

    Layout expected:
               col_labels ...
      row_label  matrix ...
      row_label  matrix ...
    """
    # Read col headers in anchor_row, starting from anchor_col+1
    col_labels = []
    col_indices = []
    for c in range(anchor_col + 1, min(sheet.max_column, anchor_col + 20) + 1):
        cell = sheet.cell(row=anchor_row, column=c)
        if cell.value is None:
            break
        col_labels.append(str(cell.value))
        col_indices.append(c)

    # Read row headers in anchor_col, starting from anchor_row+1
    row_labels = []
    row_indices = []
    for r in range(anchor_row + 1, min(sheet.max_row, anchor_row + 30) + 1):
        cell = sheet.cell(row=r, column=anchor_col)
        if cell.value is None:
            break
        row_labels.append(str(cell.value))
        row_indices.append(r)

    if len(col_labels) < 2 or len(row_labels) < 2:
        return None

    # Extract matrix
    matrix = []
    for r in row_indices:
        row_vals = []
        for c in col_indices:
            val = _to_number(sheet.cell(row=r, column=c).value)
            row_vals.append(val)
        matrix.append(row_vals)

    # Need at least 60% of cells numeric
    total_cells = len(row_indices) * len(col_indices)
    numeric_cells = sum(1 for row in matrix for v in row if v is not None)
    if numeric_cells / total_cells < 0.6:
        return None

    return {
        "rows": row_labels,
        "cols": col_labels,
        "matrix": matrix,
        "base": None,   # heuristic doesn't identify base cell
    }


# ---------------------------------------------------------------------------
# Methodology detection
# ---------------------------------------------------------------------------

def _detect_methodology(wb) -> Optional[str]:
    """Scan sheet names and first cells for methodology hints."""
    hints = [s.title.lower() for s in wb.worksheets]
    hints_all = " ".join(hints)
    for method, keywords in _METHODOLOGY_KEYWORDS.items():
        if any(k in hints_all for k in keywords):
            return method
    return None


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

_PERCENT_FIELDS = {"upside", "wacc", "growth_rate", "terminal_growth",
                   "ebitda_margin", "revenue_cagr"}


def parse_excel(path: str) -> dict:
    """Parse an Excel file and extract valuation fields heuristically.

    Returns dict with:
      extracted:    dict of {field: value} for fields found
      missing:      list of fields from REQUIRED_FIELDS that weren't found
      scenarios:    dict or None
      sensitivity:  dict or None
      sheets:       list of sheet names
    """
    extracted = {}

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        logger.error("parse_excel failed to open %s: %s", path, exc)
        return {
            "extracted": {},
            "missing": list(REQUIRED_FIELDS),
            "scenarios": None,
            "sensitivity": None,
            "sheets": [],
            "error": str(exc),
        }

    sheets = [s.title for s in wb.worksheets]

    # Methodology from sheet names
    method = _detect_methodology(wb)
    if method:
        extracted["methodology"] = method

    scenarios = None
    sensitivity = None

    for sheet in wb.worksheets:
        max_r = min(sheet.max_row or 0, 500)
        max_c = min(sheet.max_column or 0, 50)

        for row in range(1, max_r + 1):
            for col in range(1, max_c + 1):
                cell = sheet.cell(row=row, column=col)
                norm = _normalize(cell.value)
                if not norm or len(norm) > 60:
                    continue

                for field, synonyms in LABEL_MAP.items():
                    if field == "methodology":
                        # Use synonyms to detect method value in adjacent cell
                        if any(norm == s or norm == s + ":" for s in synonyms):
                            _, adj_cell = _find_adjacent_numeric(sheet, row, col, max_steps=3)
                            # actually for methodology it's a string — look for adjacent string
                            for dc in range(1, 4):
                                c2 = col + dc
                                if c2 > sheet.max_column:
                                    break
                                adj = sheet.cell(row=row, column=c2).value
                                if adj and isinstance(adj, str):
                                    norm_adj = _normalize(adj)
                                    for m, kw in _METHODOLOGY_KEYWORDS.items():
                                        if any(k in norm_adj for k in kw):
                                            extracted.setdefault(field, m)
                                            break
                                    break
                        continue

                    if field in extracted:
                        continue  # already found

                    # Exact/prefix match against synonyms
                    matched = False
                    for s in synonyms:
                        if norm == s or norm == s + ":" or norm.startswith(s + " "):
                            matched = True
                            break
                    if not matched:
                        continue

                    # Found label — get adjacent numeric cell
                    val, adj_cell = _find_adjacent_numeric(sheet, row, col, max_steps=3)
                    if val is None or adj_cell is None:
                        continue

                    if field in _PERCENT_FIELDS:
                        extracted[field] = _coerce_percent(adj_cell)
                    else:
                        num = _to_number(adj_cell.value)
                        if num is not None:
                            extracted[field] = round(num, 4)

        # Try scenarios / sensitivity on each sheet (prefer first hit)
        if scenarios is None:
            try:
                scenarios = _detect_scenarios(sheet)
            except Exception as exc:
                logger.debug("detect_scenarios skipped: %s", exc)

        if sensitivity is None:
            try:
                sensitivity = _detect_sensitivity(sheet)
            except Exception as exc:
                logger.debug("detect_sensitivity skipped: %s", exc)

    wb.close()

    missing = [f for f in REQUIRED_FIELDS if f not in extracted or extracted[f] is None]

    return {
        "extracted": extracted,
        "missing": missing,
        "scenarios": scenarios,
        "sensitivity": sensitivity,
        "sheets": sheets,
    }


# ---------------------------------------------------------------------------
# Excel → Markdown (input for Claude fallback)
# ---------------------------------------------------------------------------

def excel_to_markdown(path: str, max_cells: int = 2000) -> str:
    """Convert an xlsx file to a markdown-style representation for Claude."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        logger.error("excel_to_markdown failed to open %s: %s", path, exc)
        return ""

    parts = []
    cells_seen = 0

    for sheet in wb.worksheets:
        parts.append(f"## Sheet: {sheet.title}\n")

        # Determine usable range
        max_r = min(sheet.max_row or 0, 200)
        max_c = min(sheet.max_column or 0, 30)
        if max_r == 0 or max_c == 0:
            parts.append("(empty sheet)\n")
            continue

        # Build rows
        rows_out = []
        for row in sheet.iter_rows(min_row=1, max_row=max_r, max_col=max_c, values_only=True):
            if all(v is None for v in row):
                continue
            cells_seen += max_c
            if cells_seen > max_cells:
                break
            rows_out.append([("" if v is None else str(v)).replace("|", "\\|").replace("\n", " ")
                            for v in row])

        if not rows_out:
            parts.append("(no non-empty rows)\n")
            continue

        # Build markdown table (header row is synthetic: col letters)
        n_cols = max(len(r) for r in rows_out)
        header = ["Col " + chr(ord("A") + i) for i in range(n_cols)]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join(["---"] * n_cols) + " |")
        for r in rows_out:
            if len(r) < n_cols:
                r = r + [""] * (n_cols - len(r))
            parts.append("| " + " | ".join(r) + " |")

        if cells_seen > max_cells:
            parts.append("\n*(output truncado por limite de células)*")
            break

    wb.close()
    return "\n".join(parts)
